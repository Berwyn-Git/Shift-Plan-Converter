import tkinter as tk
from openpyxl import Workbook, load_workbook


# create a tkinter window
window = tk.Tk()

# create a label and entry for input number
label = tk.Label(window, height=35, text="Enter the total columns you need(请输入本次导入文件的总列数):\n|    |\n|    |\n|    |\n|    |\n|    |\n|    |\n|    |\n|    |\n|    |\n\  /\n\/\n\  /\n\/\n\  /\n\/\n\  /\n\/\n\  /\n\/", width=70)
label.pack()
entry = tk.Entry(window, width=35)
entry.pack()


# create a button to trigger the data changing function
def change_data():
    number = int(entry.get())
    wb = load_workbook("target_test.xlsx")
    sheet = wb.active
    for area in sheet.iter_rows(min_row=20, max_col=number):
        # change the data in the area
        pass
    wb.save("target_test_save.xlsx")


button = tk.Button(window, text="Confirm Input(确认并执行)", command=change_data)
button.pack()

# start the tkinter event loop
window.mainloop()
